from django.shortcuts import render, HttpResponse, redirect
from django.db import transaction
from django.http import JsonResponse
from student import models

from openpyxl import load_workbook

from utils.res import ResponseData


def add(request):
    class_queryset = models.Classes.objects.all()
    course_queryset = models.Course.objects.all()
    tem_dict = {'class_queryset': class_queryset, 'course_queryset': course_queryset}

    # student = models.Student.objects.get(name="xzh519")
    
    # print(student.student_detail)

    if request.method == 'POST':
        name = request.POST.get('name')
        age = request.POST.get('age', 18)  # 提供默认值
        classes_id = request.POST.get('classes')
        course_id_list = request.POST.getlist('course')
        sex = request.POST.get('sex')
        addr = request.POST.get('province')  # 获取省份

        try:
            with transaction.atomic():

                # 创建学生详细信息记录，并与学生记录关联
                detail_obj = models.StudentDetail.objects.create(
                    gender=int(sex),  # 性别转换为整数
                    addr=addr,  # 省份作为地址信息
                    phone=None,  # 假设没有提供电话
                      # 关联到刚创建的学生记录
                )
                

                # 创建学生记录
                stu_obj = models.Student.objects.create(
                    name=name, 
                    age=int(age),  # 确保年龄是整数
                    classes_id=classes_id,  # 外键关联班级
                    student_detail=detail_obj
                )
                stu_obj.save()


                # 处理选课信息
                course_list = []
                for course_id in course_id_list:
                    course_list.append(models.Student2Course(
                        student=stu_obj,
                        course_id=course_id
                    ))
                models.Student2Course.objects.bulk_create(course_list)

        except Exception as e:
            print(e)
            print('服务器错误---student/add!')
            return redirect('base:students')

        return redirect('base:students')

    return render(request, 'student/add.html', tem_dict)

class ResponseData:
    def __init__(self, status=0, message=''):
        self.status = status
        self.message = message

    def get_dict(self):
        return {'status': self.status, 'message': self.message}



def dels(request):
    current_pk = request.POST.get('current_pk')
    if current_pk:
        try:
            student_obj = models.Student.objects.get(pk=current_pk)
            student_obj.delete()
            return JsonResponse({'status': 'success'})
        except models.Student.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found.'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def edit(request, current_pk):
    stu_obj = models.Student.objects.filter(pk=current_pk).first()
    class_queryset = models.Classes.objects.all()
    course_queryset = models.Course.objects.all()
    course_id_list = [course.pk for course in stu_obj.course.all()]
    

    tem_dict = {
        'stu_obj': stu_obj,
        'class_queryset': class_queryset,
        'course_queryset': course_queryset,
        'course_id_list': course_id_list
    }
    if request.method == 'POST':
        name = request.POST.get('name')
        age = request.POST.get('age')
        classes_id = request.POST.get('classes')
        course_id_list = request.POST.getlist('course')
        sex = request.POST.get('sex')
        addr = request.POST.get('province')  # 获取省份
        print(sex)
        print(addr)
        
        try:
            with transaction.atomic():
                models.Student.objects.filter(pk=current_pk).update(name=name, age=age, classes_id=classes_id)
                models.Student2Course.objects.filter(student=stu_obj).delete()
                course_list = []
                for course_id in course_id_list:
                    course_list.append(models.Student2Course(student=stu_obj, course_id=course_id))
                models.Student2Course.objects.bulk_create(course_list)

                student_detail,created =  models.StudentDetail.objects.get_or_create(student=stu_obj)
                print(student_detail)
                student_detail.gender = sex
                student_detail.addr = addr
                student_detail.save()

        except Exception as e:
            print(e)
            print('服务器错误---student/edit!')
        finally:
            return redirect('base:students')
    return render(request, 'student/edit.html', tem_dict)


def search(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        key_word = request.POST.get('key_word')
        if category == 'name':
            ret_queryset = models.Student.objects.filter(name__contains=key_word)
        elif category == 'classes':
            ret_queryset = models.Student.objects.filter(classes__name__contains=key_word)
        elif category == 'course':
            ret_queryset = models.Student.objects.filter(course__name=key_word)
        else:
            ret_queryset = None
        return render(request, 'student/search.html', {'ret_queryset': ret_queryset, 'key_word': key_word})


def import_student(request):
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            return HttpResponse('请先选择文件')
        wb = load_workbook(import_file)
        sheet = wb.worksheets[0]
        gender_dict = {
            '保密': 0,
            '男': 1,
            '女': 2,
        }
        for cells in sheet.iter_rows():
            name = cells[0].value or None
            if name == '姓名':
                continue
            gender = cells[1].value
            gender_val = gender_dict.get(gender, 0)
            age = cells[2].value or None
            addr = cells[3].value or None
            phone = cells[4].value or None
            classes_name = cells[5].value or None
            try:
                with transaction.atomic():
                    class_obj = models.Classes.objects.filter(name=classes_name).first()
                    student_detail_obj = models.StudentDetail.objects.create(gender=gender_val, addr=addr, phone=phone)
                    models.Student.objects.create(name=name, age=age, classes=class_obj,
                                                  student_detail=student_detail_obj)
            except Exception as e:
                print(e)
                print('服务器错误---student/import_student!')
        return redirect('base:students')
